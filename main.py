from aiogram import Bot, Dispatcher, executor, types
from datetime import datetime, time, timedelta
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.dispatcher import FSMContext
from oauth2client.service_account import ServiceAccountCredentials

import keyboards as kb
import database as db
from dotenv import load_dotenv
import os
import asyncio
from send_daily_reminders import send_daily_reminders
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from functools import partial
from database import sheet  # Импортируем объект sheet
import pandas_platej as pn
from apscheduler.triggers.interval import IntervalTrigger
import openpyxl
import masis_bot as mb
from gspread import Cell
from aiogram.utils import executor
import gspread
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from aiogram.types import ParseMode
from database import get_storage

storage = get_storage()

storage.create_request(user_id, request_type, payload)

load_dotenv()
bot = Bot('')
dp = Dispatcher(bot=bot, storage=storage)
dp.middleware.setup(LoggingMiddleware())

async def on_startup(_):
    await db.db_start()
    print('Бот успешно запущен!')



class ProblemCommentYes(StatesGroup):
    row_id = State()
    comment = State()

class NewOrder(StatesGroup):
    type = State()
    name = State()
    desc = State()
    comment = State()
    partner_pagination = State()  # Новое состояние для пагинации контрагентов


class ProblemComment(StatesGroup):
    row_id = State()
    comment = State()

class Pay_reestr(StatesGroup):
    payment_type = State()
    partner = State()



# Список для хранения вопросов и ответов
questions_and_answers = []
ADMIN_IDS = []  # Список ID администраторов

# Настройки Google Sheets
scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_name('google_credentials.json', scope)
client = gspread.authorize(creds)
sheet = client.open_by_url('https://docs.google.com/spreadsheets/d/').worksheet('Секции (вид работ)')

# ВАЖНО: Убедитесь, что этот список ТОЧНО соответствует заголовкам в вашей таблице Google Sheets, включая регистр букв
expected_headers = ['Проект','Регион', 'ИНН', 'Партнер', 'Вид работ', 'Интегральный показатель', 'Рейтинг ОДК МО', 'Рейтинг ОДК МСК', 'Численность', 'sum январь', 'sum ферваль', 'sum март', 'sum апрель', 'sum май', 'sum июнь', 'sum июль', 'sum август', 'sum сентябрь', 'sum октябрь', 'sum ноябрь', 'sum декабрь']

# Список разрешённых ID пользователей
AUTHORIZED_USER_IDS = [ ]  # Замените на реальные ID пользователей





# Шаги для FSM
class AdminStates(StatesGroup):
    waiting_for_user_ids = State()
    waiting_for_message_or_file = State()


# Команда для начала процесса
@dp.message_handler(commands=['ask'], state='*')
async def ask_user_ids(message: types.Message):
    if message.from_user.id in ADMIN_IDS:  # Проверка, является ли пользователь администратором
        await AdminStates.waiting_for_user_ids.set()
        await message.reply("Введите ID пользователей, которым хотите отправить сообщение (разделитель запятая):")
    else:
        await message.reply("Вы не имеете прав для использования этой команды.")


# Получение ID пользователей
@dp.message_handler(state=AdminStates.waiting_for_user_ids)
async def get_user_ids(message: types.Message, state: FSMContext):
    try:
        user_ids = [int(user_id.strip()) for user_id in message.text.split(',')]
        await state.update_data(user_ids=user_ids)
        await AdminStates.waiting_for_message_or_file.set()
        await message.reply("Введите ваше сообщение и/или прикрепите файл:")
    except ValueError:
        await message.reply("Пожалуйста, введите корректные ID пользователей (только числа, разделенные запятыми).")


# Обработка текста и/или файла
@dp.message_handler(state=AdminStates.waiting_for_message_or_file, content_types=types.ContentType.ANY)
async def send_message_or_file(message: types.Message, state: FSMContext):
    user_data = await state.get_data()
    user_ids = user_data['user_ids']

    # Отправка текста, если он есть
    if message.text:
        for user_id in user_ids:
            try:
                await bot.send_message(user_id, message.text)
            except Exception as e:
                await message.reply(f"Не удалось отправить сообщение пользователю {user_id}: {e}")

    # Отправка файла, если он есть
    if message.document:
        for user_id in user_ids:
            try:
                await bot.send_document(user_id, message.document.file_id)
            except Exception as e:
                await message.reply(f"Не удалось отправить файл пользователю {user_id}: {e}")

    await message.reply("Сообщения и/или файлы отправлены всем пользователям.")
    await state.finish()






# Обработчик текста "Отмена"
@dp.message_handler(text='Отмена', state='*')
async def cancel_handler(message: types.Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state is not None:
        await state.finish()
        await message.answer('Действие отменено. Вы вернулись в главное меню.', reply_markup=kb.main)


@dp.callback_query_handler(text='cancel', state='*')
async def cancel_handler(callback_query: types.CallbackQuery, state: FSMContext):
    await state.finish()
    await bot.send_message(callback_query.from_user.id, 'Действие отменено. Вы вернулись в главное меню.',
                           reply_markup=kb.main)


@dp.message_handler(commands=['start'])
async def cmd_start(message: types.Message):
    user_id = message.from_user.id
    first_name = message.from_user.first_name
    last_name = message.from_user.last_name
    username = message.from_user.username

    await db.cmd_start_db(user_id, first_name, username, last_name)
    await message.answer(
        f'{first_name}, Добро пожаловать в чат Бот DEMO_CORP! ✈️✈️✈️\n'
        'Посмотрите это обучающее видео: ',
        reply_markup=kb.main
    )
    if user_id in ADMIN_IDS:
        await message.answer('Вы авторизовались как администратор!', reply_markup=kb.admin)


# Обработчики команд
@dp.message_handler(text=['Допуски'])
async def send_welcome(message: types.Message):
    await message.reply("Выберите вид работ:", reply_markup=kb.create_work_type_keyboard())

@dp.message_handler(lambda message: message.text in kb.construction_guide.keys())
async def process_work_type(message: types.Message):
    work_type = message.text
    await message.reply(f"Вы выбрали: {work_type}\nТеперь выберите параметр контроля:",
                        reply_markup=kb.create_control_param_keyboard(work_type))


@dp.message_handler(lambda message: any(message.text in params for params in kb.construction_guide.values()))
async def process_control_param(message: types.Message):
    control_param = message.text
    work_type = None
    for wt, params in kb.construction_guide.items():
        if control_param in params:
            work_type = wt
            break

    rows = mb.find_rows(work_type, control_param)
    if rows:
        response_message = "\n\n".join(
            [f" {row[0]}, \n Ссылка на нормативную документацию: {row[1]}" for row in rows])
    else:
        response_message = "Нет данных по заданным критериям."

    await message.reply(response_message)

@dp.message_handler(commands=['id'])
async def cmd_id(message: types.Message):
    await message.answer(f'{message.from_user.id}')


@dp.message_handler(text='Реестр платежей')
async def catalog(message: types.Message):
    await message.answer('Выберите тип платежа', reply_markup=kb.payment_keyboard)
    await Pay_reestr.payment_type.set()

@dp.message_handler(state=Pay_reestr.payment_type)
async def choose_payment_type(message: types.Message, state: FSMContext):
    payment_type = message.text
    async with state.proxy() as data:
        data['payment_type'] = payment_type
    await message.answer("Выберите контрагента:", reply_markup=kb.create_catalog_all_partners())
    await Pay_reestr.partner.set()



@dp.message_handler(state=Pay_reestr.partner)
async def choose_partner(message: types.Message, state: FSMContext):
    partner = message.text
    async with state.proxy() as data:
        data['partner'] = partner

    try:
        output_filename = pn.generate_and_send_report(data['payment_type'], partner)
        await message.answer_document(types.InputFile(output_filename), reply_markup=kb.main)
    except Exception as e:
        await message.answer(f'Произошла ошибка: {e}', reply_markup=kb.main)

    # Сброс состояния
    await state.finish()



@dp.message_handler(text='Главное меню')
async def cart(message: types.Message):
    user_id = message.from_user.id
    if user_id in ADMIN_IDS:
        await message.answer('Вы авторизовались как администратор!', reply_markup=kb.admin)
    else:
        await message.answer('Главное меню', reply_markup=kb.main)


@dp.message_handler(text='Контакты')
async def contacts(message: types.Message):
    await message.answer('Вопросы по Работе Бота: \n'
                         'Обратная связь для партнеров: ')


@dp.message_handler(text='Инструкции')
async def instructions(message: types.Message):
    await message.answer('Выберите инструкцию', reply_markup=kb.instructions)


@dp.message_handler(text='Видео обучение')
async def video_training(message: types.Message):
    await message.answer('Выберите видео', reply_markup=kb.video_training)


@dp.message_handler(text='Путь подрядчика')
async def contractor_guide(message: types.Message):
    await message.answer_document(open('docs/contractor_guide.pdf', 'rb'))


@dp.message_handler(text='Инструкция для Партнеров')
async def partner_guide(message: types.Message):
    await message.answer_document(open('docs/partner_guide.pdf', 'rb'))



@dp.message_handler(text='Регистрация сотрудника в системах DEMO_CORP')
async def employee_registration_guide(message: types.Message):
    await message.answer_document(open('docs/employee_registration_guide.pdf', 'rb'))

@dp.message_handler(text='Формы документов (ПТО)')
async def instructions(message: types.Message):
    await message.answer( 'Выберите тип документа из списка ниже (КЛАВИАТУРА В ПАНЕЛИ НИЖЕ):\n\n' 
                          '1. Акт на доп работы/Акт необходимости — ФОРМА DEMO_CORP\n' 
                          '2. Коммерческое предложение — В свободной форме Подрядчика\n' 
                          '3. Калькуляция — ФОРМА DEMO_CORP\n' 
                          '4. Акт сверки объем работ (АСОР) — ФОРМА DEMO_CORP\n' 
                          '5. Дефектная ведомость — ФОРМА DEMO_CORP\n' 
                          '6. Реестр ИД\n' 
                          '7. Спецификация к ДКП — ФОРМА DEMO_CORP\n' 
                          '8. Компенсация к КС-2 — ФОРМА DEMO_CORP\n' 
                          '9. Корректировочная КС-2 — ФОРМА DEMO_CORP', reply_markup=kb.form_pto )



@dp.message_handler(text='Акт на доп работы')
async def send_act_na_dop_rabot(message: types.Message):
    with open('exel/new/Акт_на_доп_работы_Форма_ПАО_ГК_Самолёт.xlsx', 'rb') as file:
        await message.answer_document(file)

@dp.message_handler(text='Акт необходимости')
async def send_akt_neobxodimosti(message: types.Message):
    with open('exel/new/Акт_необходимости_Форма_ПАО_ГК_Самолёт.xlsx', 'rb') as file:
        await message.answer_document(file)

@dp.message_handler(text='Акт сверки объем работ (АСОР)')
async def send_asor(message: types.Message):
    with open('exel/АСОР.xlsx', 'rb') as file:
        await message.answer_document(file)

@dp.message_handler(text='Корректировочная КС-2')
async def send_korr_ks2(message: types.Message):
    with open('exel/Корректировочная КС - форма.xls', 'rb') as file:
        await message.answer_document(file)

@dp.message_handler(text='Калькуляция')
async def send_kalkulyatsiya(message: types.Message):
    with open('exel/Калькуляция_образец_для_расценок.xlsx', 'rb') as file:
        await message.answer_document(file)

@dp.message_handler(text='Дефектная ведомость')
async def send_defektovaya_vedomost(message: types.Message):
    with open('exel/Дефектовочная ведомость.xlsx', 'rb') as file:
        await message.answer_document(file)

@dp.message_handler(text='Реестр ИД')
async def send_reestr_id(message: types.Message):
    with open('exel/Реестр ИД.xls', 'rb') as file:
        await message.answer_document(file)

@dp.message_handler(text='Спецификация к ДКП')
async def send_specifikatsiya_k_dkp(message: types.Message):
    with open('exel/Спецификация к ДКП.xlsx', 'rb') as file:
        await message.answer_document(file)

@dp.message_handler(text='Компенсация к КС-2 (По материалу)')
async def send_kompensatsiya_k_ks2(message: types.Message):
    with open('exel/new/Расчет компенсации по материалу.xlsm', 'rb') as file:
        await message.answer_document(file)


@dp.message_handler(text='Регистрация в S.Pro 1ч')
async def video_1(message: types.Message):
    await message.answer('Ссылка на видео 1: ')


@dp.message_handler(text='Регистрация в S.Pro 2ч')
async def video_2(message: types.Message):
    await message.answer('Ссылка на видео 2: ')


@dp.message_handler(text='Обзор меню')
async def video_1(message: types.Message):
    await message.answer('Ссылка на видео 3: ')


@dp.message_handler(text='Создание заявки')
async def video_2(message: types.Message):
    await message.answer('Ссылка на видео 4: ')


@dp.message_handler(text='Формирование РП')
async def video_1(message: types.Message):
    await message.answer('Ссылка на видео 5: ')


@dp.message_handler(text='Арматурный процесс')
async def video_2(message: types.Message):
    await message.answer('Ссылка на видео 6: ')


@dp.message_handler(text='Арматурный процесс и обычный на схемах')
async def video_1(message: types.Message):
    await message.answer('Ссылка на видео 7: ')


@dp.message_handler(text='Анулирование, редактирование заявки')
async def video_2(message: types.Message):
    await message.answer('Ссылка на видео 8: ')


@dp.message_handler(text='Материал аналог')
async def video_1(message: types.Message):
    await message.answer('Ссылка на видео 9: ')


@dp.message_handler(text='Поиск группы работ к материалу')
async def video_2(message: types.Message):
    await message.answer('Ссылка на видео 10: ')


@dp.message_handler(text='Выбрать объект')
async def add_item(message: types.Message):
    await NewOrder.type.set()
    await message.answer('Выберите Объект 🏢', reply_markup=kb.catalog_list)


@dp.message_handler(state=NewOrder.type)
async def add_item_type(message: types.Message, state: FSMContext):
    # Сохраняем выбранный тип (текст кнопки)
    async with state.proxy() as data:
        data['type'] = message.text  # Так как это обычное сообщение, а не callback

    # Переходим к следующему шагу и отправляем новую клавиатуру с проблематикой
    await message.answer('Укажите проблематику 🏗', reply_markup=kb.catalog_list_problem)
    await NewOrder.next()


@dp.message_handler(state=NewOrder.name)
async def add_item_name(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['name'] = message.text
    await NewOrder.partner_pagination.set()
    await message.answer('Выберите контрагента 🧰', reply_markup=kb.create_catalog_all_partners(page=0))


@dp.message_handler(state=NewOrder.name)
async def add_item_name(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['name'] = message.text
    await NewOrder.partner_pagination.set()
    await message.answer('Выберите контрагента 🧰', reply_markup=kb.create_catalog_all_partners(page=0))


# Обработчик пагинации контрагентов
@dp.message_handler(lambda message: message.text.startswith('Page_'), state=NewOrder.partner_pagination)
async def process_callback_kb_page(message: types.Message, state: FSMContext):
    page = int(message.text.split('_')[1])
    await message.answer("Выберите контрагента 🧰", reply_markup=kb.create_catalog_all_partners(page=page))


# Обработчик выбора контрагента
@dp.message_handler(lambda message: not message.text.startswith('Page_'), state=NewOrder.partner_pagination)
async def select_partner(message: types.Message, state: FSMContext):
    partner = message.text.strip()

    # Проверяем, что партнер есть в списке
    if partner not in kb.all_partners:
        await message.answer("Пожалуйста, выберите контрагента из доступных вариантов на клавиатуре.")
        return

    # Сохраняем выбранного партнера
    async with state.proxy() as data:
        data['desc'] = partner

    # Переход к следующему шагу FSM
    await NewOrder.comment.set()
    await message.answer(
        'Опишите проблематику ⚒ (ОБЯЗАТЕЛЬНО УКАЖИТЕ ДОМ И ВИД РАБОТ!)',
        reply_markup=kb.cancel
    )





@dp.message_handler(state=NewOrder.comment)
async def add_item_comment(message: types.Message, state: FSMContext):
    async with state.proxy() as data:
        data['comment'] = message.text
    user_id = message.from_user.id
    user_name = message.from_user.username
    await db.add_item(state, user_id, user_name)
    await message.answer('Обращение успешно создано! 🎉 Ориентировочный срок решения обращения 5 дней.',
                         reply_markup=kb.main)
    await state.finish()





@dp.message_handler(lambda message: True)
async def get_contractor_info(message: types.Message):
    # Проверка, если пользователь не в списке разрешённых
    if message.from_user.id not in AUTHORIZED_USER_IDS:
        await message.reply("У вас нет доступа к этому функционалу. Обратитесь к администраторам (Контакты)")
        return

    if sheet is None:
        await message.reply("Не удалось подключиться к Google Sheets. Проверьте настройки.")
        return

    query = message.text.strip().lower()  # Приводим запрос к нижнему регистру
    try:
        contractors = sheet.get_all_records(expected_headers=expected_headers)
        if not contractors:
            await message.reply("Данные из Google Sheets не получены.")
            return
    except gspread.exceptions.GSpreadException as e:
        await message.reply(f"Ошибка при чтении таблицы: {e}")
        return

    def safe_float(value):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    found_contractors = []

    for contractor in contractors:
        if (
                str(contractor.get('ИНН', '')).lower() == query
                or query in str(contractor.get('Партнер', '')).lower()
                or query in str(contractor.get('Регион', '')).lower()
                or query in str(contractor.get('Вид работ', '')).lower()
        ):
            found_contractors.append(contractor)

    if found_contractors:
        # Общая информация по подрядчику
        inn = found_contractors[0].get('ИНН', 'Не указан')
        partner = found_contractors[0].get('Партнер', 'Не указан')

        # Сбор и вывод значений рейтингов для отладки
        odk_mo_raw = [c.get('Рейтинг ОДК МО') for c in found_contractors]
        odk_msk_raw = [c.get('Рейтинг ОДК МСК') for c in found_contractors]

        print(message.from_user.id, "Рейтинг ОДК МО (сырые):", odk_mo_raw)
        print(message.from_user.id, "Рейтинг ОДК МСК (сырые):", odk_msk_raw)

        odk_mo_list = [
            safe_float(value) for value in odk_mo_raw
            if safe_float(value) is not None
        ]
        odk_msk_list = [
            safe_float(value) for value in odk_msk_raw
            if safe_float(value) is not None
        ]

        print("Рейтинг ОДК МО (float):", odk_mo_list)
        print("Рейтинг ОДК МСК (float):", odk_msk_list)

        # Расчёт средних
        avg_odk_mo = sum(odk_mo_list) / len(odk_mo_list) if odk_mo_list else 0
        avg_odk_msk = sum(odk_msk_list) / len(odk_msk_list) if odk_msk_list else 0

        avg_odk_mo_str = "{0:.2f}".format(avg_odk_mo / 100).replace('.', ',')
        avg_odk_msk_str = "{0:.2f}".format(avg_odk_msk / 100).replace('.', ',')

        response = f"""ИНН: {inn}
    Подрядчик: {partner}
    Средний рейтинг ОДК МО: {avg_odk_mo_str}
    Средний рейтинг ОДК МСК: {avg_odk_msk_str}

    """

        for contractor in found_contractors:
            project = contractor.get('Проект', 'Не указан')
            work_type = contractor.get('Вид работ', 'Не указан')
            region = contractor.get('Регион', 'Не указан')
            count = contractor.get('Численность', 'Не указана')

            # Интегральный показатель для текущей записи
            try:
                integral = float(contractor.get('Интегральный показатель', 0)) / 100
                integral_str = f"{integral:.2f}".replace('.', ',')
            except:
                integral_str = str(contractor.get('Интегральный показатель', '—'))

            response += f"""Проект: {project}
    Лот: {work_type}
    Регион: {region}
    Интегральный показатель: {integral_str}
    Численность: {count}
    """

            months_map = [
                ('sum март', 'мар'), ('sum апрель', 'апр'), ('sum май', 'май'), ('sum июнь', 'июн'),
                ('sum июль', 'июл'), ('sum август', 'авг'), ('sum сентябрь', 'сен'),
                ('sum октябрь', 'окт'), ('sum ноябрь', 'ноя'), ('sum декабрь', 'дек'),
            ]

            header_row = ''
            value_row = ''

            for field_name, short_name in months_map:
                value = contractor.get(field_name)
                if value in (None, '', '—') or not str(value).isdigit():
                    value_str = '0'
                else:
                    value_str = str(value)

                # Месяц без пробелов
                header_row += f'{short_name}|'

                # Значения: выравниваем по инструкции
                if len(value_str) == 1:
                    value_row += f'  {value_str}  |'
                elif len(value_str) == 2:
                    value_row += f' {value_str}  |'
                else:
                    value_row += f'{value_str}|'

            month_info = f"```\nСекции по месяцам:\n{header_row}\n{value_row}\n```"

            response += month_info + "\n---\n"

        # Отправка с разбивкой
        while len(response) > 4096:
            await message.reply(response[:4096])
            response = response[4096:]

        await message.reply(response)
    else:
        await handle_unprocessed_message(message)






async def handle_unprocessed_message(message: types.Message):
    user_id = message.from_user.id
    username = message.from_user.username or "No username"
    text = message.text

    # Логирование необработанного сообщения
    log_message = f"Необработанное сообщение от {username} ({user_id}): {text}"
    print(log_message)  # Вывод в консоль для отслеживания

    # Запись вопроса в Google Sheets (можно записывать только необработанные сообщения)
    await db.record_question_and_answer(user_id, username, text)  # Убедитесь, что db.record_question_and_answer - асинхронная функция

    # Отправка необработанного сообщения всем администраторам
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(admin_id, f"⚠️ Необработанное сообщение от {username} ({user_id}):\n{text}", parse_mode=ParseMode.HTML)
        except Exception as e:
            print(f"Не удалось отправить необработанное сообщение администратору {admin_id}: {e}")

    # Ответ пользователю о том, что сообщение принято (можно сделать более информативным)
    await message.answer(
        'Данного контрагента или команды нет в списке. \nВаше сообщение принято и будет рассмотрено администратором.'
    )











# Регистрация хэндлеров для обработчиков из send_daily_reminders.py
@dp.callback_query_handler(lambda call: call.data.startswith('resolve_'))
async def resolve_problem(call: types.CallbackQuery, state: FSMContext):
    row_id = int(call.data.split('_')[1])

    # Сохраняем row_id в состоянии для последующего получения комментария
    await state.update_data(row_id=row_id)

    # Запрашиваем комментарий от пользователя
    await call.message.edit_text("Пожалуйста, укажите комментарий для решения проблемы:")
    await ProblemCommentYes.comment.set()
    await call.answer("Укажите комментарий для решения проблемы.")


@dp.message_handler(state=ProblemCommentYes.comment)
async def add_solution_comment(message: types.Message, state: FSMContext):
    data = await state.get_data()
    row_id = data['row_id']
    comment = message.text

    # Обновляем Google Sheets
    sheet.update_cell(row_id, 10, comment)  # Предположим, что колонка "Комментарий" - 10-я
    sheet.update_cell(row_id, 12, 'Да')  # Обновляем колонку "Решено?" на "Да"
    sheet.update_cell(row_id, 14, 'Да')  # Обновляем колонку "Ответ получен?" на "Да"

    await message.answer('Проблема отмечена как решенная. Ваш комментарий сохранен.', reply_markup=kb.main)
    await state.finish()

    # Отправка уведомления инициатору проблемы
    user_id = sheet.cell(row_id, 1).value  # Получаем ID инициатора
    date_reciev = sheet.cell(row_id, 7).value  # Получаем дату обращения
    table_problem = sheet.cell(row_id, 3).value  # Описание проблемы

    if user_id:
        await bot.send_message(chat_id=user_id,
                               text=f"Ваше обращение от {date_reciev}, по проблеме '{table_problem}' решено.\nКомментарий: {comment}")
    else:
        print("ID инициатора не указан. Не удалось отправить сообщение.")

    await send_daily_reminders(bot)  # Продолжаем отправлять напоминания


# Регистрация хэндлеров для обработчиков из send_daily_reminders.py
@dp.callback_query_handler(lambda call: call.data.startswith('resolve2_'))
async def resolve2_problem(call: types.CallbackQuery):
    row_id = int(call.data.split('_')[1])
    sheet.update_cell(row_id, 13, 'Да')  # Предположим, что колонка "Решено?" - это третья колонка
    await call.answer("Проблема отмечена как решенная.")
    user_id = sheet.cell(row_id, 9).value  # Получаем ID инициатора проблемы
    date_reciev = sheet.cell(row_id, 7).value  #
    table_problem = sheet.cell(row_id, 3).value  # Получаем описание проблемы
    if user_id:
        await bot.send_messagehat_id=user_id,
                               sheet=(cf"Обращение от {date_reciev}, по проблематике {table_problem} решено, Контрагент подтвердил.")
    else:
        print("ID инициатора не указан. Не удалось отправить сообщение.")
    await send_daily_reminders(bot)  # Продолжаем отправлять напоминания


@dp.callback_query_handler(lambda call: call.data.startswith('no2_'))
async def problem_not_resolved2(call: types.CallbackQuery, state: FSMContext):
    row_id = int(call.data.split('_')[1])
    sheet.update_cell(row_id, 13, 'Нет')
    user_id = sheet.cell(row_id, 9).value  # Получаем ID инициатора проблемы
    date_reciev = sheet.cell(row_id, 7).value  #
    table_problem = sheet.cell(row_id, 3).value  # Получаем описание проблемы
    if user_id:
        await bot.send_message(chat_id=user_id,
                               text=f"Обращение от {date_reciev}, по проблематике {table_problem} не решено, Контрагент не подтвердил.")
    else:
        print("ID инициатора не указан. Не удалось отправить сообщение.")
    await send_daily_reminders(bot)

@dp.callback_query_handler(lambda call: call.data.startswith('no_'))
async def problem_not_resolved(call: types.CallbackQuery, state: FSMContext):
    row_id = int(call.data.split('_')[1])
    await state.update_data(row_id=row_id)
    await call.message.edit_text("\nПожалуйста, укажите причину, почему проблема не решена:")
    await ProblemComment.comment.set()
    await call.answer("Укажите комментарий для проблемы.")


@dp.message_handler(state=ProblemComment.comment)
async def add_problem_comment(message: types.Message, state: FSMContext):
    data = await state.get_data()
    row_id = data['row_id']
    comment = message.text
    sheet.update_cell(row_id, 10, comment)  # Предположим, что колонка "Комментарий по проблеме" - это 17 колонка
    sheet.update_cell(row_id, 12, 'Нет')  # Обновляем колонку "Решено?" на "Нет"
    sheet.update_cell(row_id, 14, 'Да')  # Обновляем колонку "Ответ получен?" на "Да"

    await message.answer('Ваш комментарий записан. Проблема отмечена как не решенная.', reply_markup=kb.main)
    await state.finish()

    # Отправка уведомления инициатору проблемы
    user_id = sheet.cell(row_id, 1).value  # Получаем ID инициатора проблемы
    date_reciev = sheet.cell(row_id, 7).value  # Получаем дату получения обращения
    table_problem = sheet.cell(row_id, 3).value  # Получаем описание проблемы

    if user_id:
        await bot.send_message(chat_id=user_id,
                               text=f"Ваше обращение от {date_reciev}, по проблематике {table_problem} не решено. Комментарий: {comment}")
    else:
        print("ID инициатора не указан. Не удалось отправить сообщение.")

    await send_daily_reminders(bot)  # Продолжаем отправлять напоминания




async def check_user_activity():
    """
    Функция для проверки активности пользователей и обновления информации в Google Sheets.
    """
    directory = db.spreadsheet.worksheet("Справочник")
    users = directory.get_all_values()  # Извлекаем все данные за один запрос

    # Определяем индекс столбца "Активность" для обновления
    header = users[0]
    activity_col = header.index('Активность') + 1  # +1 для корректного индекса в Google Sheets

    updates = []  # Список для хранения обновлений активности

    for idx, row in enumerate(users[1:], start=2):  # Начинаем с 2, так как 1-я строка заголовки
        user_id = row[header.index('ID')]
        activity_status = 'Нет'  # Устанавливаем статус по умолчанию как 'Нет'

        try:
            # Отправляем сообщение пользователю
            sent_message = await bot.send_message(user_id, "Проверка активности бота.")
            activity_status = 'Да'  # Если сообщение отправлено успешно, активность "Да"

            # Удаляем сообщение сразу после отправки
            await bot.delete_message(chat_id=user_id, message_id=sent_message.message_id)
        except Exception as e:
            print(f"Ошибка при проверке активности для пользователя {user_id}: {e}")
            activity_status = 'Нет'  # Если произошла ошибка, активность "Нет"

        # Добавляем данные для обновления
        updates.append((idx, activity_col, activity_status))
        await asyncio.sleep(0.1)  # Минимальная пауза для избегания перегрузки

    # Обновляем данные в Google Sheets
    try:
        cell_updates = [Cell(row, col, value) for row, col, value in updates]
        directory.update_cells(cell_updates)
        print("Проверка активности завершена и данные обновлены.")
    except gspread.exceptions.APIError as e:
        print(f"APIError: {e}")


# Функция для проверки времени и выполнения задач
async def daily_task_wrapper(_=None):
    now = datetime.now()
    print(now.time())
    current_time = now.time() # для теста время + 10 часов (now + timedelta(hours=10)).time()
    start_time = time(10, 0)  # 10:00 AM
    end_time = time(20, 0)  # 8:00 PM

    print(f"Текущее время: {now.strftime('%Y-%m-%d %H:%M:%S')}")

    if start_time <= current_time <= end_time:
        await send_daily_reminders(bot)
    else:
        print("Сейчас не время для отправки сообщений. Ожидание до следующего разрешенного интервала.")


# Функция для планировщика
async def scheduler():
    # Создание планировщика
    scheduler = AsyncIOScheduler()

    # Функция для проверки времени и выполнения задачи


    # Установка задачи на отправку сообщения каждую минуту
    scheduler.add_job(daily_task_wrapper, IntervalTrigger(minutes=360)) # 360
    scheduler.add_job(check_user_activity, 'interval', minutes=4320) # 4320

    # Запуск планировщика
    scheduler.start()

    try:
        while True:
            await asyncio.sleep(60)  # Пауза на 60 секунд
    except (KeyboardInterrupt, SystemExit):
        scheduler.shutdown()






if __name__ == "__main__":
    loop = asyncio.get_event_loop()
    loop.create_task(scheduler())
    executor.start_polling(dp, skip_updates=True, on_startup=on_startup)
